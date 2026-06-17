from io import BytesIO

from openpyxl import load_workbook

from web.app import REPORTS, create_app


def _metadata(nullable: str) -> bytes:
    return f"""<?xml version="1.0"?>
<edmx:Edmx xmlns:edmx="http://schemas.microsoft.com/ado/2007/06/edmx"
 xmlns="http://schemas.microsoft.com/ado/2008/09/edm">
  <edmx:DataServices>
    <Schema Namespace="SFOData">
      <EntityType Name="EmpJob">
        <Property Name="department" Type="Edm.String" Nullable="{nullable}" MaxLength="32"/>
      </EntityType>
    </Schema>
  </edmx:DataServices>
</edmx:Edmx>""".encode()


def test_web_upload_generates_results_and_excel_download() -> None:
    REPORTS.clear()
    app = create_app()
    app.config["TESTING"] = True
    client = app.test_client()

    response = client.post(
        "/compare",
        data={
            "before_label": "Baseline",
            "after_label": "Changed",
            "before_files": (BytesIO(_metadata("true")), "before.xml"),
            "after_files": (BytesIO(_metadata("false")), "after.xml"),
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    assert b"Comparison complete" in response.data
    assert b"CRITICAL" in response.data

    report_id = next(iter(REPORTS))
    download = client.get(f"/download/{report_id}/xlsx")
    assert download.status_code == 200
    workbook = load_workbook(BytesIO(download.data))
    assert "Changes" in workbook.sheetnames


def test_web_requires_both_snapshots() -> None:
    app = create_app()
    app.config["TESTING"] = True
    client = app.test_client()

    response = client.post(
        "/compare",
        data={"before_files": (BytesIO(_metadata("true")), "before.xml")},
        content_type="multipart/form-data",
    )

    assert response.status_code == 400
    assert b"Choose at least one After file" in response.data
