from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from sf_change_ledger.diff import compare_snapshots
from sf_change_ledger.ingest import load_snapshot
from sf_change_ledger.report import render_excel, write_report
from sf_change_ledger.risk import assess_diff

ROOT = Path(__file__).parents[1]


def _result():
    before = load_snapshot(ROOT / "samples" / "before", "Before")
    after = load_snapshot(ROOT / "samples" / "after", "After")
    return assess_diff(compare_snapshots(before, after))


def test_excel_contains_expected_worksheets() -> None:
    workbook = load_workbook(BytesIO(render_excel(_result())))

    assert workbook.sheetnames == [
        "Summary",
        "Changes",
        "Property Diffs",
        "Test Checklist",
    ]
    assert workbook["Summary"]["B6"].value == 5
    assert workbook["Changes"]["A2"].value == "CRITICAL"


def test_write_report_supports_xlsx(tmp_path: Path) -> None:
    out = tmp_path / "report.xlsx"
    write_report(_result(), out)

    assert out.exists()
    assert out.stat().st_size > 1000
