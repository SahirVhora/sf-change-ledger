from __future__ import annotations

import html
import json
from dataclasses import asdict
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from sf_change_ledger.models import DiffResult


def write_report(result: DiffResult, out: str | Path) -> None:
    path = Path(out)
    path.parent.mkdir(parents=True, exist_ok=True)
    suffix = path.suffix.lower()
    if suffix == ".json":
        path.write_text(render_json(result), encoding="utf-8")
    elif suffix == ".xlsx":
        path.write_bytes(render_excel(result))
    elif suffix in {".html", ".htm"}:
        path.write_text(render_html(result), encoding="utf-8")
    else:
        path.write_text(render_markdown(result), encoding="utf-8")


def render_json(result: DiffResult) -> str:
    return json.dumps(asdict(result), indent=2, default=str)


def render_excel(result: DiffResult) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _write_summary_sheet(summary, result)
    _write_changes_sheet(workbook.create_sheet("Changes"), result)
    _write_property_sheet(workbook.create_sheet("Property Diffs"), result)
    _write_testing_sheet(workbook.create_sheet("Test Checklist"), result)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_summary_sheet(sheet, result: DiffResult) -> None:
    sheet.append(["SF Change Ledger", "Configuration change report"])
    sheet.append(["Baseline", result.left_label])
    sheet.append(["Comparison", result.right_label])
    sheet.append([])
    sheet.append(["Metric", "Count"])
    metrics = [
        ("Total changes", len(result.changes)),
        ("Added", len(result.added)),
        ("Removed", len(result.removed)),
        ("Modified", len(result.modified)),
        ("Critical", result.by_severity.get("CRITICAL", 0)),
        ("High", result.by_severity.get("HIGH", 0)),
        ("Medium", result.by_severity.get("MEDIUM", 0)),
        ("Low", result.by_severity.get("LOW", 0)),
    ]
    for metric in metrics:
        sheet.append(metric)
    sheet["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    sheet["B1"].font = Font(size=12, bold=True, color="FFFFFF")
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="17324D")
    _style_header(sheet, 5)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 42


def _write_changes_sheet(sheet, result: DiffResult) -> None:
    headers = [
        "Severity",
        "Change",
        "Object Type",
        "Object",
        "Object ID",
        "Why It Matters",
        "Test Focus",
    ]
    sheet.append(headers)
    for change in result.changes:
        sheet.append(
            [
                change.severity,
                change.kind.value,
                change.object_type,
                change.label,
                change.object_id,
                change.explanation,
                "\n".join(change.test_focus),
            ]
        )
    _style_header(sheet, 1)
    _apply_severity_fills(sheet, 1)
    _format_table_sheet(sheet, [12, 12, 20, 32, 46, 72, 58])


def _write_property_sheet(sheet, result: DiffResult) -> None:
    sheet.append(["Severity", "Object", "Property", "Before", "After"])
    for change in result.changes:
        for prop in change.property_changes:
            sheet.append(
                [
                    change.severity,
                    change.label,
                    prop.path,
                    _display_value(prop.before),
                    _display_value(prop.after),
                ]
            )
    _style_header(sheet, 1)
    _apply_severity_fills(sheet, 1)
    _format_table_sheet(sheet, [12, 34, 42, 42, 42])


def _write_testing_sheet(sheet, result: DiffResult) -> None:
    sheet.append(["Status", "Test Activity", "Related Severity"])
    rows: dict[str, set[str]] = {}
    for change in result.changes:
        for test in change.test_focus:
            rows.setdefault(test, set()).add(change.severity)
    for test, severities in sorted(rows.items()):
        ordered = sorted(
            severities, key=lambda item: ["CRITICAL", "HIGH", "MEDIUM", "LOW"].index(item)
        )
        sheet.append(["Not started", test, ", ".join(ordered)])
    _style_header(sheet, 1)
    _format_table_sheet(sheet, [16, 86, 22])


def _style_header(sheet, row: int) -> None:
    for cell in sheet[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="245C6A")
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = f"A{row + 1}"
    sheet.auto_filter.ref = sheet.dimensions


def _apply_severity_fills(sheet, column: int) -> None:
    colors = {
        "CRITICAL": "FECACA",
        "HIGH": "FED7AA",
        "MEDIUM": "FEF3C7",
        "LOW": "D1FAE5",
    }
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=column)
        fill = colors.get(str(cell.value))
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(bold=True)


def _format_table_sheet(sheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _display_value(value) -> str:
    if value is None:
        return "(not set)"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=True, sort_keys=True)
    return str(value)


def render_markdown(result: DiffResult) -> str:
    lines = [
        "# SF Change Ledger Report",
        "",
        f"Compared `{result.left_label}` to `{result.right_label}`.",
        "",
        "## Summary",
        "",
        f"- Total changes: {len(result.changes)}",
        f"- Added: {len(result.added)}",
        f"- Removed: {len(result.removed)}",
        f"- Modified: {len(result.modified)}",
        f"- Critical: {result.by_severity.get('CRITICAL', 0)}",
        f"- High: {result.by_severity.get('HIGH', 0)}",
        f"- Medium: {result.by_severity.get('MEDIUM', 0)}",
        f"- Low: {result.by_severity.get('LOW', 0)}",
        "",
        "## Testing Checklist",
        "",
    ]
    tests = sorted({test for change in result.changes for test in change.test_focus})
    if tests:
        lines.extend(f"- [ ] {test}" for test in tests)
    else:
        lines.append("- [ ] No changes detected.")

    lines.extend(["", "## High-Risk Changes", ""])
    risky = [change for change in result.changes if change.severity in {"CRITICAL", "HIGH"}]
    if not risky:
        lines.append("No critical or high-risk changes found.")
    for change in risky:
        lines.extend(_render_change_markdown(change))

    lines.extend(["", "## Detailed Changes", ""])
    if not result.changes:
        lines.append("No semantic configuration changes found.")
    for change in result.changes:
        lines.extend(_render_change_markdown(change))

    return "\n".join(lines) + "\n"


def _render_change_markdown(change) -> list[str]:
    lines = [
        f"### {change.severity}: {change.label}",
        "",
        f"- Type: `{change.object_type}`",
        f"- Change: `{change.kind.value}`",
        f"- ID: `{change.object_id}`",
        f"- Why it matters: {change.explanation}",
    ]
    if change.property_changes:
        lines.extend(["", "| Property | Before | After |", "|---|---|---|"])
        for prop in change.property_changes[:20]:
            lines.append(f"| `{prop.path}` | `{prop.before}` | `{prop.after}` |")
        if len(change.property_changes) > 20:
            lines.append(f"| ... | {len(change.property_changes) - 20} more changes | |")
    lines.append("")
    return lines


def render_html(result: DiffResult) -> str:
    rows = []
    for change in result.changes:
        property_rows = "".join(
            f"<tr><td>{html.escape(prop.path)}</td><td>{html.escape(_display_value(prop.before))}</td>"
            f"<td>{html.escape(_display_value(prop.after))}</td></tr>"
            for prop in change.property_changes
        )
        detail = (
            f"<table><thead><tr><th>Property</th><th>Before</th><th>After</th></tr></thead>"
            f"<tbody>{property_rows}</tbody></table>"
            if property_rows
            else '<p class="muted">No property-level details for this change.</p>'
        )
        rows.append(
            f"""<section class="change">
  <div class="change-head">
    <span class="severity {change.severity.lower()}">{html.escape(change.severity)}</span>
    <div><h2>{html.escape(change.label)}</h2><p>{html.escape(change.kind.value)} · {html.escape(change.object_type)}</p></div>
  </div>
  <p>{html.escape(change.explanation)}</p>
  {detail}
</section>"""
        )
    changes_html = "\n".join(rows) or "<p>No semantic configuration changes found.</p>"
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>SF Change Ledger Report</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{ font-family: Arial, sans-serif; margin: 0; line-height: 1.5; color: #172033; background: #f4f7f8; }}
    header {{ background: #17324d; color: white; padding: 30px max(24px, calc((100vw - 1120px) / 2)); }}
    header h1 {{ margin: 0 0 4px; font-size: 28px; }}
    header p {{ margin: 0; color: #c9d8e5; }}
    main {{ max-width: 1120px; margin: 0 auto; padding: 26px 24px 48px; }}
    .summary {{ display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); border: 1px solid #d6e0e4; background: white; margin-bottom: 22px; }}
    .metric {{ padding: 18px; border-right: 1px solid #d6e0e4; }}
    .metric:last-child {{ border-right: 0; }}
    .metric strong {{ display: block; font-size: 24px; }}
    .metric span {{ color: #60717c; font-size: 13px; }}
    .change {{ background: white; border: 1px solid #d6e0e4; padding: 20px; margin-bottom: 14px; }}
    .change-head {{ display: flex; gap: 12px; align-items: flex-start; }}
    .change h2 {{ margin: 0; font-size: 18px; }}
    .change-head p {{ margin: 2px 0 0; color: #60717c; font-size: 13px; }}
    .severity {{ padding: 4px 8px; font-size: 11px; font-weight: 700; }}
    .critical {{ background: #fecaca; color: #991b1b; }}
    .high {{ background: #fed7aa; color: #9a3412; }}
    .medium {{ background: #fef3c7; color: #854d0e; }}
    .low {{ background: #d1fae5; color: #065f46; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 14px; }}
    th, td {{ border: 1px solid #d6e0e4; padding: 8px 10px; text-align: left; vertical-align: top; }}
    th {{ background: #eef3f4; font-size: 12px; }}
    .muted {{ color: #60717c; }}
    @media (max-width: 720px) {{
      .summary {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
      .metric:nth-child(2) {{ border-right: 0; }}
      table {{ font-size: 12px; }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>SF Change Ledger</h1>
    <p>{html.escape(result.left_label)} compared with {html.escape(result.right_label)}</p>
  </header>
  <main>
    <section class="summary">
      <div class="metric"><strong>{len(result.changes)}</strong><span>Total changes</span></div>
      <div class="metric"><strong>{result.by_severity.get("CRITICAL", 0)}</strong><span>Critical</span></div>
      <div class="metric"><strong>{result.by_severity.get("HIGH", 0)}</strong><span>High</span></div>
      <div class="metric"><strong>{len(result.modified)}</strong><span>Modified</span></div>
    </section>
    {changes_html}
  </main>
</body>
</html>
"""
